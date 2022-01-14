# -*- coding: utf-8 -*-
import time
import requests
from lxml import etree
import json

headers = {
            'user-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.71 Safari/537.36 Edg/97.0.1072.55',
            'accept-encoding': 'gzip, deflate, br',
            'authority': 'search.jd.com',
            'accept':'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'accept-language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
            'cookie': ''
        }


import openpyxl
outwb = openpyxl.Workbook()
outws = outwb.create_sheet(index=0)

outws.cell(row=1,column=1,value="index")
outws.cell(row=1,column=2,value="title")
outws.cell(row=1,column=3,value="price")
outws.cell(row=1,column=4,value="CommentCount")


#commentcount("71929438514")
###获取每一页的商品数据
def getlist(url):
    global  count
    #url="https://search.jd.com/search?keyword=笔记本&wq=笔记本&ev=exbrand_联想&page=9&s=241&click=1"
    res = requests.get(url,headers=headers)
    res.encoding = 'utf-8'
    text = res.text


    selector = etree.HTML(text)
    list = selector.xpath('//*[@id="J_goodsList"]/ul/li')
    count = 1
    product_request_links = []
    for i in list:
        title=i.xpath('.//div[@class="p-name p-name-type-2"]/a/em/text()')[0]
        price = i.xpath('.//div[@class="p-price"]/strong/i/text()')[0]
        product_id = i.xpath('.//div[@class="p-commit"]/strong/a/@id')[0].replace("J_comment_","")
        product_request_link = "https:" + i.xpath('.//div[@class="p-name p-name-type-2"]/a/@href')[0]
        shop_name = i.xpath('.//div[@class="p-shop"]/span/a/@title')[0]
        #comment_count = commentcount(product_id)
        print(title)
        print(price)
        print(shop_name)
        #print(comment_count)

        outws.cell(row=count, column=1, value=str(count-1))
        outws.cell(row=count, column=2, value=str(title))
        outws.cell(row=count, column=3, value=str(price))
        outws.cell(row=count, column=4, value=str(shop_name))
        #outws.cell(row=count, column=4, value=str(comment_count))
        product_request_links.append(product_request_link)
        count = count +1
        #print("-----")

    for link in product_request_links:
        req = requests.get(link,headers=headers)
        req.encoding = 'utf-8'
        text = req.text
        
        selector = etree.HTML(text)
        item_list = selector.xpath('//*[@id="choose-attr-1"]/div[@class="dd"]/div') # div[@class="dd"]/div[@class="item"]
        for item in item_list:
            product_data_value = item.xpath(".//@data-value")[0]
            data_sku = item.xpath(".//@data-sku")[0]
            data_text = item.xpath(".//@title")
            temp_link = "https://item.jd.com/" + data_sku + ".html#crumb-wrap"
            requ = requests.get(temp_link, headers=headers)            
            requ.encoding = 'utf-8'
            time.sleep(0.2)
            temp_text = requ.text

            selector = etree.HTML(temp_text)
            price_spc_product_class = selector.xpath('//*[@class="summary-price J-summary-price"]/div[@class="dd"]/*[@class="p-price"]')[0]
            price_spc_product = price_spc_product_class.xpath('.//span')[0].text
            dfdfdfdfd = price_spc_product_class.xpath('.//span[contains(@class, "price")]')
            print(price_spc_product)
        print(len(item_list))



#遍历每一页
def getpage():
    page=1
    s = 1
    goods_name = '充电宝'
    brand_name = '华为'
    for i in range(1,6):
        print("page="+str(page)+",s="+str(s))
        url = "https://search.jd.com/Search?keyword=" + goods_name + "%20" + brand_name + "&enc=utf-8&wq=" + goods_name + "%20" + brand_name
        getlist(url)
        page = page+2
        s = s+60



#开始爬取
getpage()
#getlist()
filename = '充电宝'
outwb.save("京东商品-" + filename + ".xls")#保存